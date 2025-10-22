# learning/service/file_service.py
import openpyxl
from typing import List, Dict, Tuple
import logging

logger = logging.getLogger(__name__)


class ExcelQuizImporter:
    """Service to handle Excel file imports for quiz creation"""

    REQUIRED_COLUMNS = ['question', 'answer', 'results']

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None

    def load_workbook(self) -> bool:
        """Load and validate Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            logger.info(f"Successfully loaded workbook: {self.file_path}")
            return True
        except FileNotFoundError:
            logger.error(f"File not found: {self.file_path}")
            raise ValueError(f"Excel file not found at {self.file_path}")
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            raise ValueError(f"Failed to load Excel file: {str(e)}")

    def validate_headers(self) -> bool:
        """Validate that Excel has required columns"""
        headers = [cell.value.lower().strip() if cell.value else ''
                   for cell in self.worksheet[1]]

        for required_col in self.REQUIRED_COLUMNS:
            if required_col.lower() not in headers:
                raise ValueError(f"Missing required column: '{required_col}'")

        return True

    def get_column_indices(self) -> Dict[str, int]:
        """Get column indices for required columns"""
        headers = [cell.value.lower().strip() if cell.value else ''
                   for cell in self.worksheet[1]]

        indices = {}
        for required_col in self.REQUIRED_COLUMNS:
            try:
                indices[required_col] = headers.index(required_col.lower()) + 1
            except ValueError:
                raise ValueError(f"Column '{required_col}' not found")

        return indices

    def parse_quiz_data(self) -> List[Dict[str, any]]:
        """
        Parse Excel file and extract quiz data

        Returns:
            List of dictionaries with keys: question_text, answer_options (list)
            where each answer_option has: option_text, is_correct
        """
        try:
            self.load_workbook()
            self.validate_headers()
            col_indices = self.get_column_indices()

            quiz_data = []

            # Skip header row (row 1)
            for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    question_cell = row[col_indices['question'] - 1]
                    answer_cell = row[col_indices['answer'] - 1]
                    result_cell = row[col_indices['results'] - 1]

                    question_text = question_cell.value
                    answer_text = answer_cell.value
                    result_text = result_cell.value

                    # Skip empty rows
                    if not question_text or not answer_text:
                        logger.warning(f"Skipping row {row_idx}: missing question or answer")
                        continue

                    # Convert result to boolean (handles: true/false, 1/0, yes/no, etc.)
                    is_correct = self._parse_boolean(result_text)

                    # Check if question already exists in quiz_data
                    existing_question = next(
                        (q for q in quiz_data if q['question_text'] == str(question_text).strip()),
                        None
                    )

                    answer_option = {
                        'option_text': str(answer_text).strip(),
                        'is_correct': is_correct
                    }

                    if existing_question:
                        # Add answer option to existing question
                        existing_question['answer_options'].append(answer_option)
                    else:
                        # Create new question entry
                        quiz_data.append({
                            'question_text': str(question_text).strip(),
                            'answer_options': [answer_option]
                        })

                except Exception as e:
                    logger.error(f"Error parsing row {row_idx}: {str(e)}")
                    raise ValueError(f"Error parsing row {row_idx}: {str(e)}")

            if not quiz_data:
                raise ValueError("No valid quiz data found in Excel file")

            # Validate that each question has at least one correct answer
            for question in quiz_data:
                if not any(opt['is_correct'] for opt in question['answer_options']):
                    raise ValueError(
                        f"Question '{question['question_text']}' has no correct answer"
                    )

            logger.info(f"Successfully parsed {len(quiz_data)} questions from Excel")
            return quiz_data

        finally:
            if self.workbook:
                self.workbook.close()

    @staticmethod
    def _parse_boolean(value) -> bool:
        """Convert various boolean representations to boolean"""
        if value is None:
            return False

        value_str = str(value).lower().strip()

        true_values = ['true', '1', 'yes', 'y', 'correct']
        false_values = ['false', '0', 'no', 'n', 'incorrect']

        if value_str in true_values:
            return True
        elif value_str in false_values:
            return False
        else:
            raise ValueError(f"Cannot parse '{value}' as boolean. Use 'true'/'false', '1'/'0', 'yes'/'no', etc.")